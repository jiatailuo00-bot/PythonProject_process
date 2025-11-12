from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

from ..models import ScriptMetadata, ScriptParameter, ScriptRunResponse
from .base import ScriptDefinition

CUSTOMER_PATTERN = re.compile(r"\[客户\]")
LINE_START_PATTERN = re.compile(r"^\s*\[[^\]]+\]")


def _resolve_column_index(header_row, target_name: str) -> Optional[int]:
    for idx, cell in enumerate(header_row, start=1):
        value = str(cell.value).strip() if cell.value is not None else ""
        if value == target_name:
            return idx
    return None


def _run(params: Dict) -> ScriptRunResponse:
    excel_path = Path(params.get("excel_path", "")).expanduser()
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel文件不存在: {excel_path}")

    sheet_name = params.get("sheet_name") or None
    context_column = (params.get("context_column") or "最终传参上下文").strip()
    latest_column = (params.get("latest_customer_column") or "最新客户消息").strip()
    if not context_column:
        context_column = "最终传参上下文"
    if not latest_column:
        latest_column = "最新客户消息"
    output_path_param = params.get("output_path")
    output_filename = params.get("output_filename")

    if output_path_param:
        candidate = Path(output_path_param).expanduser()
        if candidate.is_dir() or not candidate.suffix:
            # treat as directory; append filename
            filename = output_filename or excel_path.name
            output_path = candidate / filename
        else:
            output_path = candidate
    else:
        filename = output_filename or excel_path.name
        output_path = excel_path.parent / filename

    output_path = output_path.expanduser()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    context_idx = _resolve_column_index(header_row, context_column)
    latest_idx = _resolve_column_index(header_row, latest_column)

    if context_idx is None:
        raise ValueError(f"未找到列: {context_column}")
    if latest_idx is None:
        raise ValueError(f"未找到列: {latest_column}")

    updates = 0
    processed = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        processed += 1
        ctx_cell = row[context_idx - 1]
        latest_cell = row[latest_idx - 1]
        ctx_val = ctx_cell.value
        if not isinstance(ctx_val, str):
            continue

        lines = ctx_val.splitlines()
        idx = 0
        last_customer_line = None

        while idx < len(lines):
            line = lines[idx]
            stripped = line.strip()
            if stripped and CUSTOMER_PATTERN.search(stripped):
                segment = [stripped]
                idx += 1
                while idx < len(lines):
                    next_line = lines[idx]
                    if LINE_START_PATTERN.match(next_line.strip()):
                        break
                    clean = next_line.strip()
                    if clean:
                        segment.append(clean)
                    idx += 1
                last_customer_line = "\n".join(segment).strip()
                continue
            idx += 1

        if last_customer_line and latest_cell.value != last_customer_line:
            latest_cell.value = last_customer_line
            updates += 1

    wb.save(output_path)

    message = f"处理完成：共扫描 {processed} 行，更新 {updates} 行"
    data = {
        "rows_processed": processed,
        "rows_updated": updates,
        "output_file": str(output_path),
    }
    return ScriptRunResponse(success=True, message=message, data=data)


SCRIPT_DEFINITION = ScriptDefinition(
    metadata=ScriptMetadata(
        id="update_latest_customer_message",
        name="同步最新客户消息",
        description="从'最终传参上下文'中提取最新的客户话术并覆盖'最新客户消息'列",
        category="Excel处理",
        parameters=[
            ScriptParameter(
                name="excel_path",
                label="Excel文件路径",
                type="path",
                description="需要处理的xlsx路径",
                example="uploads/修改最后客户消息.xlsx",
            ),
            ScriptParameter(
                name="sheet_name",
                label="工作表名称",
                type="string",
                required=False,
                description="若留空则默认使用活动工作表",
            ),
            ScriptParameter(
                name="context_column",
                label="上下文列名",
                type="string",
                required=False,
                description="包含完整历史对话的列，默认'最终传参上下文'",
                example="最终传参上下文",
            ),
            ScriptParameter(
                name="latest_customer_column",
                label="最新客户列名",
                type="string",
                required=False,
                description="需要被覆盖的列，默认'最新客户消息'",
                example="最新客户消息",
            ),
            ScriptParameter(
                name="output_path",
                label="输出文件路径",
                type="path",
                required=False,
                description="可填写新的保存路径；留空则覆盖原文件",
            ),
            ScriptParameter(
                name="output_filename",
                label="输出文件名",
                type="string",
                required=False,
                description="若未指定则沿用原文件名",
            ),
        ],
        output_description="返回统计信息和输出文件路径",
    ),
    runner=_run,
)
